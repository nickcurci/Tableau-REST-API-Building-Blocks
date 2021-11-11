'''
    req = s.get(
        f'https://tableau.website.org/api/3.11/sites/siteid/views/{Provider_Detail_ID}/data?vf_Market={Market}&maxAge=1',
        headers=headers)

    market_path = path + "/" + Market + "/" + Market + " Provider Detail Report - " + yesterday.strftime(
        '%Y %m %d') + ".xlsx"
'''

def get_Market_Excel(Market, index):
    print('Trying excel function')

    # Initializing variables needed for report creation
    # the folder which all reports generated will be stored in (Ex. 2021 Jun)
    # the folder which seperates reports to specific hospitals (Ex. GSMC)
    MarketFolder = provmarket_list["Market"][index]

    # Sending request for KPI report using the provider filter and name to grab the individual report
    print('sending request')

    req = s.get(
        f'https://tableau.website.org/api/3.11/sites/siteid/views/{Provider_Detail_ID}/data?vf_Market={Market}&maxAge=1',
        headers=headers)
    req.raise_for_status()
    print('request successful')

    market_path = path + "/" + Market + "/" + Market + " Provider Detail Report - " + yesterday.strftime(
        '%Y %m %d') + ".xlsx"
    shutil.copy(template, market_path)
    print(f'setting market path = {market_path}')
    print('copying pivot table template')
    print('decoding api response data')
    reqdata = StringIO(req.content.decode())
    print('sending it to CSV')
    df = pd.read_csv(reqdata)
    print('filerting out unwanted ALL values')
    df_filtered = df[~df['Market'].str.contains('All')]
    print('getting rid of commas')
    df_filtered['Measure Values'] = df_filtered['Measure Values'].str.replace(',', '').astype(float)
    reqdata.close()
    print('response closed and setting writer')

    writer = pd.ExcelWriter(market_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    print('writer created, sending to excel')
    df_filtered.to_excel(writer, sheet_name='Provider Detail', index=False)
    writer.save()
    writer.close()

# This needs to be done to remove the extra sheet created
    from openpyxl import load_workbook
    wb = load_workbook(market_path)
    if 'Provider Detail1' in wb.sheetnames:
        wb.remove(wb['Provider Detail'])
        wb.save(market_path)

        ss_sheet = wb['Provider Detail1']
        ss_sheet.title = 'Provider Detail'
        wb.save(market_path)
    else:
        print('nothing to be alarmed about')
        wb.save(market_path)
    print('sent to excel and saved')

    print('importing win32com')
    import win32com.client as win32
    print('dispatching excel application')
    xlapp = win32.DispatchEx('Excel.Application')
    print('setting display alerts off')
    xlapp.DisplayAlerts = False
    print('setting visibility off')
    xlapp.Visible = False
    print(f'opening workbook at {market_path}')
    try:
        xlbook = xlapp.Workbooks.Open(market_path)
    except:
        print('breaking the law')
        pass
    print('refreshing workbook')
    # Refresh all pivot tables
    try:
        time.sleep(5)
        xlbook.RefreshAll()
        xlapp.CalculateUntilAsyncQueriesDone()
    except:
        print('breaking the law)')
    print('saving')
    try:
        xlbook.Save()
    except:
        print('saving failed but yolo')
        pass
    print('closing')
    xlbook.Close()
    print('quitting')
    xlapp.Quit()
    # Make sure Excel completely closes
    print('deleting')
    del xlbook
    del xlapp
    print('closing all excel interations opened')
    print(f'{Market} report generatd in {str(market_path)}')

for Market in market_list:
    get_Market_Excel(Market, i)